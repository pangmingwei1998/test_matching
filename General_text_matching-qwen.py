#!/usr/bin/env python3
"""
通用条款匹配脚本
功能：语义向量召回 + LLM精判 + Excel导出
适用于任意两个责任标准文档的比对
"""

import json
import os
import time
from typing import List, Dict, Tuple, Any
import requests
import numpy as np
import pandas as pd
import openai
from tqdm import tqdm

import faiss
import torch
from transformers import AutoTokenizer, AutoModel, AutoModelForSequenceClassification

# 配置 CUDA 内存管理以减少显存碎片
os.environ['PYTORCH_CUDA_ALLOC_CONF'] = 'expandable_segments:True'


# ==================== 配置参数 ====================
class Config:
    # 文件路径
    A_FILE = "/home/pmw/h20/Text_matching/RBA_A.json"
    B_FILE = "/home/pmw/h20/Text_matching/Apple_standard.json"
    OUTPUT_EXCEL = "/home/pmw/h20/Text_matching/General_matching_results.xlsx"
    OUTPUT_HTML = "/home/pmw/h20/Text_matching/General_matching_results.html"

    # Qwen3-Embedding-8B 嵌入模型 (Qwen/Qwen3-Embedding-8B)
    # 支持本地路径或 Hugging Face 模型名
    EMBEDDING_MODEL = "Qwen/Qwen3-Embedding-8B"
    # 8B 模型约 16GB，使用 CPU 运行（稳定但较慢）
    EMBEDDING_DEVICE = "cpu"

    # 检查本地缓存是否存在
    HF_CACHE = os.path.expanduser("~/.cache/huggingface/hub/models--Qwen--Qwen3-Embedding-8B")
    USE_LOCAL_ONLY = os.path.exists(HF_CACHE)  # 如果缓存存在，强制使用本地模式

    # BGE-Reranker-Base 模型
    RERANKER_MODEL = "BAAI/bge-reranker-base"  # BGE Reranker 模型
    RERANKER_DEVICE = "cpu"  # Reranker 使用 CPU，节省显存
    RERANKER_TOP_K = 5  # Rerank 后取 Top-K
    ENABLE_RERANKER = True  # Reranker 开关：True=启用，False=禁用

    # 向量检索参数
    TOP_K = 20  # 召回Top-K候选（增大召回数量）

    # LLM API 配置
    LLM_API_BASE = "http://10.71.5.24:8000/v1"
    LLM_API_KEY = "empty"  # 本地服务通常不需要key
    LLM_MODEL = "gpt-3.5-turbo"
    LLM_TIMEOUT = 60
    LLM_MAX_RETRIES = 3

    # 相似度阈值
    SIMILARITY_THRESHOLD = 0.8  # 低于此分数的匹配不进行LLM判断


# ==================== 工具函数 ====================
def load_json_documents(file_path: str) -> List[Dict[str, Any]]:
    """加载JSON格式的文档，支持两种格式：
    1. JSON数组格式：[{...}, {...}, ...]
    2. 每行一个JSON对象：{...}\n{...}\n...
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read().strip()

    # 尝试方式1: JSON数组格式
    try:
        documents = json.loads(content)
        if isinstance(documents, list):
            print(f"  检测到JSON数组格式，加载 {len(documents)} 条记录")
            return documents
    except json.JSONDecodeError:
        pass

    # 尝试方式2: 每行一个JSON对象
    documents = []
    lines = content.split('\n')
    line_num = 0
    while line_num < len(lines):
        line = lines[line_num].strip()
        if not line:
            line_num += 1
            continue

        try:
            # 尝试单行解析
            doc = json.loads(line)
            documents.append(doc)
            line_num += 1
        except json.JSONDecodeError:
            # 尝试多行解析（找到完整的JSON对象）
            json_str = line
            nested_line = line_num + 1
            while nested_line < len(lines):
                next_line = lines[nested_line]
                json_str += '\n' + next_line
                try:
                    doc = json.loads(json_str)
                    documents.append(doc)
                    line_num = nested_line + 1
                    break
                except json.JSONDecodeError:
                    nested_line += 1
            else:
                print(f"警告: 第{line_num + 1}行JSON解析失败（尝试了多行合并）")
                line_num += 1

    print(f"  检测到每行JSON对象格式，加载 {len(documents)} 条记录")
    return documents


def filter_content_blocks(documents: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """过滤出只包含 content 字段的文档块（排除 Preamble）"""
    content_docs = []
    for doc in documents:
        # 只保留包含 'content' 字段的块
        if 'content' in doc:
            content_docs.append(doc)
    return content_docs


def build_embedding_text(doc: Dict[str, Any]) -> str:
    """构建用于 embedding 的文本，包含层级路径信息

    格式: 每个层级独占一行，最后是内容
    例如:
    Anti-Discrimination
    Supplier Code of Conduct
    Supplier Responsibility Standards
    1. Policy
    具体内容...
    """
    parts = []

    # 添加层级信息（每个层级单独一行）
    if doc.get('Theme'):
        parts.append(doc['Theme'])

    level_1 = doc.get('level_1', {})
    if level_1 and level_1.get('title'):
        title = level_1['title']
        if level_1.get('id'):
            parts.append(f"{level_1['id']}. {title}")
        else:
            parts.append(title)

    level_2 = doc.get('level_2', {})
    if level_2 and level_2.get('title'):
        title = level_2['title']
        if level_2.get('id'):
            parts.append(f"{level_2['id']}. {title}")
        else:
            parts.append(title)

    level_3 = doc.get('level_3', {})
    if level_3 and level_3.get('title'):
        title = level_3['title']
        if level_3.get('id'):
            parts.append(f"{level_3['id']}. {title}")
        else:
            parts.append(title)

    # 添加实际内容
    content = doc.get('content', '')
    if content:
        parts.append(content)

    return '\n'.join(parts)


def truncate_text(text: str, max_length: int = 512) -> str:
    """截断过长的文本"""
    if len(text) <= max_length:
        return text
    return text[:max_length] + "..."


# ==================== Qwen 向量嵌入 ====================
class QwenEmbedder:
    """使用 Qwen3-Embedding-8B 模型生成文本嵌入向量

    Qwen Embedding 模型不需要前缀，直接使用原始文本
    """

    def __init__(self, model_name: str = Config.EMBEDDING_MODEL, device: str = Config.EMBEDDING_DEVICE):
        print(f"正在加载 Qwen3-Embedding-8B 模型 ({device}模式)...")
        self.device = device

        # 禁用 huggingface_hub 的网络检查
        os.environ['HF_HUB_DISABLE_TELEMETRY'] = '1'

        # 获取本地模型路径
        if Config.USE_LOCAL_ONLY:
            # 查找实际的 snapshot 路径
            import glob
            snapshot_pattern = os.path.expanduser("~/.cache/huggingface/hub/models--Qwen--Qwen3-Embedding-8B/snapshots/*")
            snapshot_dirs = glob.glob(snapshot_pattern)
            if snapshot_dirs:
                # 找到包含完整模型文件的 snapshot
                valid_snapshot = None
                for snapshot_dir in snapshot_dirs:
                    if os.path.exists(os.path.join(snapshot_dir, "config.json")) and \
                       os.path.exists(os.path.join(snapshot_dir, "tokenizer.json")):
                        valid_snapshot = snapshot_dir
                        break

                if valid_snapshot:
                    print(f"  使用本地缓存: {valid_snapshot}")
                    model_to_load = valid_snapshot
                    use_local = True
                else:
                    print(f"  本地缓存不完整，尝试从远程下载: {model_name}")
                    model_to_load = model_name
                    use_local = False
            else:
                print(f"  本地缓存未找到，尝试从远程下载: {model_name}")
                model_to_load = model_name
                use_local = False
        else:
            print(f"  使用远程模型: {model_name}")
            model_to_load = model_name
            use_local = False

        load_kwargs = {
            "local_files_only": use_local,
            "trust_remote_code": True
        }

        try:
            self.tokenizer = AutoTokenizer.from_pretrained(model_to_load, **load_kwargs)
            self.model = AutoModel.from_pretrained(model_to_load, **load_kwargs)
            self.model.to(device)
            self.model.eval()
            print("Qwen3-Embedding-8B 模型加载完成")

        except Exception as e:
            print(f"\n错误: 模型加载失败: {e}")
            print("\n解决方案:")
            print("1. 清理显存: python -c \"import torch; torch.cuda.empty_cache()\"")
            print("2. 检查网络连接")
            print("3. 设置环境变量使用镜像站: export HF_ENDPOINT=https://hf-mirror.com")
            print("4. 手动下载模型: huggingface-cli download Qwen/Qwen3-Embedding-8B")
            raise

    def encode_queries(self, texts: List[str], batch_size: int = 32) -> np.ndarray:
        """将查询文本编码为向量"""
        return self._encode(texts, batch_size)

    def encode_passages(self, texts: List[str], batch_size: int = 32) -> np.ndarray:
        """将文档文本编码为向量"""
        return self._encode(texts, batch_size)

    def encode(self, texts: List[str], batch_size: int = 32, is_query: bool = False) -> np.ndarray:
        """将文本列表编码为向量

        Args:
            texts: 文本列表
            batch_size: 批次大小
            is_query: Qwen Embedding 不区分 query/passage，此参数保留兼容性
        """
        return self._encode(texts, batch_size)

    def _encode(self, texts: List[str], batch_size: int = 32) -> np.ndarray:
        """内部编码方法"""
        all_embeddings = []
        total_texts = len(texts)

        for i in range(0, total_texts, batch_size):
            batch_texts = texts[i:i + batch_size]
            current_batch_num = i // batch_size + 1
            total_batches = (total_texts + batch_size - 1) // batch_size

            print(f"  正在处理批次 {current_batch_num}/{total_batches} ({len(batch_texts)} 条文本)...")

            # Tokenize
            encoded_input = self.tokenizer(
                batch_texts,
                padding=True,
                truncation=True,
                max_length=512,
                return_tensors='pt'
            )

            # Move to device
            encoded_input = {k: v.to(self.device) for k, v in encoded_input.items()}

            # Encode
            with torch.no_grad():
                model_output = self.model(**encoded_input)
                # Qwen 使用平均池化
                embeddings = self._mean_pooling(model_output, encoded_input['attention_mask'])
                # 归一化
                embeddings = torch.nn.functional.normalize(embeddings, p=2, dim=1)

            # 立即移到 CPU 并释放 GPU 显存
            all_embeddings.append(embeddings.cpu().numpy())

            # 清理中间变量
            del embeddings, encoded_input, model_output
            if self.device == "cuda":
                torch.cuda.empty_cache()

        return np.vstack(all_embeddings)

    def _mean_pooling(self, model_output, attention_mask):
        """平均池化"""
        token_embeddings = model_output[0]
        input_mask_expanded = attention_mask.unsqueeze(-1).expand(token_embeddings.size()).float()
        return torch.sum(token_embeddings * input_mask_expanded, 1) / torch.clamp(input_mask_expanded.sum(1), min=1e-9)


# ==================== 向量索引 ====================
class VectorIndex:
    """使用 FAISS 构建向量索引"""

    def __init__(self, embeddings: np.ndarray):
        dimension = embeddings.shape[1]
        self.index = faiss.IndexFlatIP(dimension)  # 内积相似度（向量已归一化）
        self.index.add(embeddings.astype('float32'))
        self.dimension = dimension

    def search(self, query_embedding: np.ndarray, top_k: int = 5) -> Tuple[np.ndarray, np.ndarray]:
        """搜索最相似的向量
        返回: (相似度分数, 索引)
        """
        similarities, indices = self.index.search(query_embedding.astype('float32'), top_k)
        return similarities, indices


# ==================== Reranker 重排序 ====================
class BGEReranker:
    """使用 BGE-Reranker 模型进行重排序"""

    def __init__(self, model_name: str = Config.RERANKER_MODEL, device: str = Config.RERANKER_DEVICE):
        print(f"正在加载 BGE-Reranker 模型 ({device}模式)...")
        self.device = device

        # 禁用 telemetry
        os.environ['HF_HUB_DISABLE_TELEMETRY'] = '1'

        # 检查本地缓存
        import glob
        reranker_cache_pattern = os.path.expanduser(f"~/.cache/huggingface/hub/models--BAAI--bge-reranker-base/snapshots/*")
        reranker_cache_dirs = glob.glob(reranker_cache_pattern)

        use_local = False
        model_to_load = model_name

        if reranker_cache_dirs:
            for snapshot_dir in reranker_cache_dirs:
                if os.path.exists(os.path.join(snapshot_dir, "config.json")):
                    print(f"  使用本地 Reranker 缓存: {snapshot_dir}")
                    model_to_load = snapshot_dir
                    use_local = True
                    break
        else:
            print(f"  首次使用 Reranker，将下载模型（约1.1GB）")

        load_kwargs = {
            "local_files_only": use_local,
            "trust_remote_code": True
        }

        try:
            self.tokenizer = AutoTokenizer.from_pretrained(model_to_load, **load_kwargs)
            # 设置 pad token（如果不存在）
            if self.tokenizer.pad_token is None:
                self.tokenizer.pad_token = self.tokenizer.eos_token
                self.tokenizer.pad_token_id = self.tokenizer.eos_token_id

            self.model = AutoModelForSequenceClassification.from_pretrained(
                model_to_load,
                **load_kwargs
            )
            self.model.to(device)
            self.model.eval()
            print("Reranker 模型加载完成")

        except Exception as e:
            print(f"\n警告: Reranker 模型加载失败: {e}")
            print("将跳过 Reranker 步骤，直接使用向量检索结果")
            self.model = None
            self.tokenizer = None

    def rerank(self, query: str, candidates: List[Dict[str, Any]], top_k: int = None) -> List[Dict[str, Any]]:
        """对候选结果进行重排序
        Args:
            query: 查询文本
            candidates: 候选列表，每个元素包含 {'text': str, 'index': int, 'score': float, ...}
            top_k: 返回前 k 个结果
        Returns:
            重排序后的候选列表
        """
        if self.model is None or not candidates:
            return candidates[:top_k] if top_k else candidates

        if top_k is None:
            top_k = Config.RERANKER_TOP_K

        # 准备输入
        texts = [c.get('content', c.get('text', '')) for c in candidates]

        # 逐个计算分数（避免 batch size > 1 的 pad_token 问题）
        rerank_scores = []

        for doc_text in texts:
            # Tokenize 单个样本
            inputs = self.tokenizer(
                [[query, doc_text]],  # 保持 list of lists 格式
                padding=True,
                truncation=True,
                max_length=512,
                return_tensors='pt'
            ).to(self.device)

            # 计算分数
            with torch.no_grad():
                outputs = self.model(**inputs)
                # 假设模型输出是 logits，取第一个作为相关性分数
                score = outputs.logits[0][0].item() if outputs.logits.dim() > 1 else outputs.logits.item()
                rerank_scores.append(score)

        # 更新分数并排序
        for i, candidate in enumerate(candidates):
            candidate['rerank_score'] = float(rerank_scores[i])

        # 按 rerank_score 降序排序
        reranked = sorted(candidates, key=lambda x: x['rerank_score'], reverse=True)

        return reranked[:top_k]

    def is_available(self) -> bool:
        """检查 Reranker 是否可用"""
        return self.model is not None


# ==================== LLM 精判 ====================
class LLMJudge:
    """使用 LLM 判断两个段落的相关性"""

    # 相关性等级
    RELEVANCE_NOT_RELATED = "不相关"
    RELEVANCE_WEAK = "弱相关"
    RELEVANCE_STRONG = "强相关"

    def __init__(self):
        self.client = openai.OpenAI(
            api_key=Config.LLM_API_KEY,
            base_url=Config.LLM_API_BASE
        )

    def judge(self, text1: str, text2: str) -> Tuple[str, str]:
        """判断两段文本的相关性
        返回: (相关性等级, 理由说明)
        """
        prompt = self._build_prompt(text1, text2)

        for attempt in range(Config.LLM_MAX_RETRIES):
            try:
                response = self.client.chat.completions.create(
                    model=Config.LLM_MODEL,
                    messages=[
                        {
                            "role": "system",
                            "content": "你是一个专业的责任标准文档分析专家。你需要判断两段文本在'责任义务层面'是否相关。"
                        },
                        {
                            "role": "user",
                            "content": prompt
                        }
                    ],
                    temperature=0.1,
                    timeout=Config.LLM_TIMEOUT
                )

                result = response.choices[0].message.content.strip()
                return self._parse_result(result)

            except Exception as e:
                if attempt < Config.LLM_MAX_RETRIES - 1:
                    wait_time = (attempt + 1) * 2
                    print(f"LLM 调用失败，{wait_time}秒后重试... 错误: {e}")
                    time.sleep(wait_time)
                else:
                    print(f"LLM 调用失败，使用默认值: {e}")
                    return self.RELEVANCE_NOT_RELATED, f"调用失败: {str(e)}"

    def _build_prompt(self, text1: str, text2: str) -> str:
        """构建 LLM 判断提示词"""
        return f"""请判断以下两段责任标准条款在"责任义务层面"是否相关。

【条款 A】：
{text1}

【条款 B】：
{text2}

从以下维度判断：
1. 是否涉及相似的责任或义务主题
2. 是否规定相似的要求或标准
3. 覆盖范围关系（完全一致/部分覆盖/互补）
4. 严格程度差异


请仅返回以下格式的结果（不要输出其他内容）：
相关性：[不相关/弱相关/强相关]
理由：["匹配类型": "两条款完全一致", "两条款部分覆盖", "两条款补充说明";
        简要说明判断理由，不超过100字]
"""

    def _parse_result(self, result: str) -> Tuple[str, str]:
        """解析 LLM 返回结果"""
        result = result.strip()

        # 提取相关性等级
        relevance = self.RELEVANCE_NOT_RELATED
        if self.RELEVANCE_STRONG in result:
            relevance = self.RELEVANCE_STRONG
        elif self.RELEVANCE_WEAK in result:
            relevance = self.RELEVANCE_WEAK
        elif self.RELEVANCE_NOT_RELATED in result:
            relevance = self.RELEVANCE_NOT_RELATED

        # 提取理由
        reason = ""
        if "理由：" in result:
            reason = result.split("理由：", 1)[1].strip()
        elif "Reason:" in result:
            reason = result.split("Reason:", 1)[1].strip()

        return relevance, reason


# ==================== 主匹配流程 ====================
class TextMatcher:
    """文本匹配主流程"""

    def __init__(self):
        print("=" * 60)
        print("通用条款匹配系统")
        print("=" * 60)

        # 加载文档
        print("\n[1/4] 加载文档...")
        all_a_docs = load_json_documents(Config.A_FILE)
        all_b_docs = load_json_documents(Config.B_FILE)

        # 只使用 content 字段，过滤 Preamble
        self.a_docs = filter_content_blocks(all_a_docs)
        self.b_docs = filter_content_blocks(all_b_docs)

        print(f"  - A文件: {len(all_a_docs)} 条（其中 content: {len(self.a_docs)} 条）")
        print(f"  - B文件: {len(all_b_docs)} 条（其中 content: {len(self.b_docs)} 条）")

        # 初始化嵌入模型
        print("\n[2/4] 初始化 Qwen3-Embedding-8B 嵌入模型...")
        self.embedder = QwenEmbedder()

        # 构建B文件文档向量索引
        print("\n[3/4] 构建B文件文档向量索引...")
        # 使用带层级路径的文本进行 embedding
        b_texts = [build_embedding_text(doc) for doc in self.b_docs]
        # B 文档作为 passage（文档库）
        b_embeddings = self.embedder.encode_passages(b_texts)
        self.vector_index = VectorIndex(b_embeddings)
        print(f"  - 向量维度: {b_embeddings.shape[1]}")
        print(f"  - 索引完成")

        # 初始化 LLM 判断器
        print("\n[4/5] 初始化 LLM 判断器...")
        self.llm_judge = LLMJudge()
        print("  - API 地址:", Config.LLM_API_BASE)
        print("  - 模型:", Config.LLM_MODEL)

        # 初始化 Reranker
        print("\n[5/5] 初始化 Reranker...")
        if Config.ENABLE_RERANKER:
            self.reranker = BGEReranker()
            if self.reranker.is_available():
                print(f"  - Reranker 已启用 (Top-K: {Config.RERANKER_TOP_K})")
            else:
                print("  - Reranker 加载失败，将跳过重排序步骤")
        else:
            self.reranker = None
            print("  - Reranker 已禁用（ENABLE_RERANKER = False）")

        print("\n" + "=" * 60)
        print("初始化完成，开始匹配...")
        print("=" * 60 + "\n")

        # 保存文档数量统计和文件名（去掉 .json 后缀）
        self.doc_counts = {
            'a_docs': len(self.a_docs),
            'b_docs': len(self.b_docs),
            'a_file_name': os.path.basename(Config.A_FILE).replace('.json', ''),
            'b_file_name': os.path.basename(Config.B_FILE).replace('.json', '')
        }

    def match(self) -> List[Dict[str, Any]]:
        """执行匹配流程：向量检索 -> Rerank -> LLM 精判"""
        results = []

        # 对A文件的每个content段落进行匹配
        for a_doc in tqdm(self.a_docs, desc="匹配进度"):
            # 使用带层级路径的文本进行向量检索
            a_text_for_embedding = build_embedding_text(a_doc)
            a_text = a_doc.get('content', '')  # 用于展示的原始内容

            # 1. 向量检索 Top-K（召回更多候选）
            # 使用带层级路径的文本进行查询
            # A 文档作为 query（查询）
            query_embedding = self.embedder.encode_queries([a_text_for_embedding])
            similarities, indices = self.vector_index.search(query_embedding, Config.TOP_K)

            # 2. 准备候选列表
            candidates = []
            for similarity, idx in zip(similarities[0], indices[0]):
                if similarity < Config.SIMILARITY_THRESHOLD * 0.5:  # 降低阈值，让 Reranker 来筛选
                    continue
                b_doc = self.b_docs[idx]
                candidates.append({
                    'doc': b_doc,
                    'content': b_doc.get('content', ''),
                    'similarity': float(similarity),
                    'index': idx
                })

            if not candidates:
                # 没有候选，添加空匹配
                results.append({
                    'A文件条款': a_text,
                    'B文件条款': '',
                    '向量相似度': '',
                    'Rerank分数': '',
                    '排名': '',
                    'LLM判断结果': '',
                    'LLM判断理由': '',
                    'B文件路径': '',
                    'A文件路径': a_doc.get('path', ''),
                })
                continue

            # 3. Rerank 重排序
            if self.reranker is not None and self.reranker.is_available():
                # 使用带层级路径的文本进行 rerank
                reranked = self.reranker.rerank(a_text_for_embedding, candidates, top_k=Config.RERANKER_TOP_K)
                top_candidates = reranked
            else:
                # Reranker 不可用或已禁用，直接使用向量相似度排序
                top_candidates = sorted(candidates, key=lambda x: x['similarity'], reverse=True)[:Config.RERANKER_TOP_K]

            # 4. 对 Top-K 结果进行 LLM 精判
            has_match = False
            for rank, candidate in enumerate(top_candidates, 1):
                b_doc = candidate['doc']
                # 获取B文档用于展示的原始内容
                b_text = b_doc.get('content', '')

                # LLM 精判（使用原始内容进行判断，不包含层级路径）
                llm_relevance, llm_reason = self.llm_judge.judge(a_text, b_text)

                # 保存结果
                result = {
                    'A文件条款': a_text,
                    'B文件条款': b_text,
                    '向量相似度': round(candidate['similarity'], 4),
                    'Rerank分数': round(candidate.get('rerank_score', 0), 4),
                    '排名': rank,
                    'LLM判断结果': llm_relevance,
                    'LLM判断理由': llm_reason,
                    'B文件路径': b_doc.get('path', ''),
                    'A文件路径': a_doc.get('path', ''),
                }
                results.append(result)
                has_match = True

            # 如果没有匹配结果，添加空匹配
            if not has_match:
                results.append({
                    'A文件条款': a_text,
                    'B文件条款': '',
                    '向量相似度': '',
                    'Rerank分数': '',
                    '排名': '',
                    'LLM判断结果': '',
                    'LLM判断理由': '',
                    'B文件路径': '',
                    'A文件路径': a_doc.get('path', ''),
                })

        return results

    def export_to_excel(self, results: List[Dict[str, Any]], output_path: str = None):
        """导出结果到 Excel，相同的"A文件条款"合并单元格"""
        if output_path is None:
            output_path = Config.OUTPUT_EXCEL

        print(f"\n正在导出结果到 {output_path}...")

        df = pd.DataFrame(results)

        # 调整列顺序
        columns_order = [
            'A文件条款',
            'B文件条款',
            '向量相似度',
            'Rerank分数',
            'LLM判断结果',
            'LLM判断理由',
            '排名',
            'A文件路径',
            'B文件路径',
        ]

        # 只保留存在的列
        columns_order = [col for col in columns_order if col in df.columns]
        df = df[columns_order]

        # 导出到 Excel
        from openpyxl.styles import Alignment, Font, Border, Side
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='匹配结果')
            worksheet = writer.sheets['匹配结果']

            # 调整列宽
            worksheet.column_dimensions['A'].width = 60  # A文件条款
            worksheet.column_dimensions['B'].width = 60  # B文件条款
            worksheet.column_dimensions['C'].width = 15  # 向量相似度
            worksheet.column_dimensions['D'].width = 15  # Rerank分数
            worksheet.column_dimensions['E'].width = 15  # LLM判断结果
            worksheet.column_dimensions['F'].width = 40  # LLM判断理由
            worksheet.column_dimensions['G'].width = 10  # 排名
            worksheet.column_dimensions['H'].width = 40  # A文件路径
            worksheet.column_dimensions['I'].width = 40  # B文件路径

            # 设置所有数据行的行高为 200
            for row in range(2, len(df) + 2):  # 从第2行开始（第1行是标题）
                worksheet.row_dimensions[row].height = 200

            # 设置标题行样式
            header_font = Font(bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for cell in worksheet[1]:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # 合并相同的A文件条款单元格
            # 从数据行开始（第2行，Excel索引为2）
            start_row = 2
            current_value = None
            merge_start_row = 2

            for row_idx in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_idx, column=1).value

                if cell_value != current_value:
                    # 如果之前的值相同且有多行，合并单元格
                    if current_value is not None and merge_start_row < row_idx - 1:
                        worksheet.merge_cells(f'A{merge_start_row}:A{row_idx - 1}')
                        # 设置合并后的单元格对齐方式
                        merged_cell = worksheet.cell(row=merge_start_row, column=1)
                        merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    current_value = cell_value
                    merge_start_row = row_idx

            # 处理最后一组相同的值
            if merge_start_row < len(df) + 2:
                worksheet.merge_cells(f'A{merge_start_row}:A{len(df) + 1}')
                merged_cell = worksheet.cell(row=merge_start_row, column=1)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # 设置所有数据行的对齐方式和边框
            for row in range(2, len(df) + 2):
                for col in range(1, len(columns_order) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = Alignment(
                        horizontal='left' if col in [1, 2, 5, 7, 8] else 'center',
                        vertical='top',
                        wrap_text=True
                    )
                    cell.border = thin_border

        print(f"导出完成！共 {len(results)} 条记录")

        # 计算统计信息
        stats = self._calculate_stats(results)
        return stats

    def _calculate_stats(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """计算结果统计信息"""
        relevance_counts = {}
        empty_match = 0
        for r in results:
            if r['LLM判断结果'] == '':
                empty_match += 1
            else:
                relevance = r['LLM判断结果']
                relevance_counts[relevance] = relevance_counts.get(relevance, 0) + 1

        return {
            'total': len(results),
            'empty_match': empty_match,
            'relevance_counts': relevance_counts
        }

    def export_to_html(self, results: List[Dict[str, Any]], stats: Dict[str, Any], doc_counts: Dict[str, int], output_path: str = None):
        """导出结果到 HTML，带美观样式和统计信息"""
        if output_path is None:
            output_path = Config.OUTPUT_HTML

        print(f"\n正在导出结果到 {output_path}...")

        df = pd.DataFrame(results)

        # 调整列顺序
        columns_order = [
            'A文件条款',
            'B文件条款',
            '向量相似度',
            'Rerank分数',
            'LLM判断结果',
            'LLM判断理由',
            '排名',
            'A文件路径',
            'B文件路径',
        ]
        columns_order = [col for col in columns_order if col in df.columns]
        df = df[columns_order]

        # 生成 HTML
        html_content = self._generate_html(df, columns_order, stats, doc_counts)

        # 写入文件
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        print(f"HTML 导出完成！文件路径: {output_path}")

    def _generate_html(self, df: pd.DataFrame, columns_order: List[str], stats: Dict[str, Any], doc_counts: Dict[str, int]) -> str:
        """生成完整的 HTML 内容"""

        # 计算合并单元格的 rowspan
        merge_spans = self._calculate_merge_spans(df)

        # 统计信息 HTML
        stats_html = self._generate_stats_html(stats, doc_counts)

        # 表格内容 HTML
        table_html = self._generate_table_html(df, columns_order, merge_spans, doc_counts)

        # 组装完整 HTML
        html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>RBA Clause Matching Results</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, "Noto Sans", sans-serif, "Apple Color Emoji", "Segoe UI Emoji", "Segoe UI Symbol", "Noto Color Emoji";
            line-height: 1.6;
            color: #333;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}

        .container {{
            max-width: 1800px;
            margin: 0 auto;
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            overflow: hidden;
        }}

        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}

        .header h1 {{
            font-size: 2.5rem;
            font-weight: 700;
            margin-bottom: 10px;
            text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.2);
        }}

        .header .subtitle {{
            font-size: 1.1rem;
            opacity: 0.9;
        }}

        .stats-section {{
            padding: 30px 40px;
            background: #f8f9fa;
            border-bottom: 2px solid #e9ecef;
        }}

        .stats-title {{
            font-size: 1.3rem;
            font-weight: 600;
            color: #495057;
            margin-bottom: 20px;
            display: flex;
            align-items: center;
        }}

        .stats-title::before {{
            content: "📊";
            margin-right: 10px;
            font-size: 1.5rem;
        }}

        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
        }}

        .stat-card {{
            background: white;
            border-radius: 12px;
            padding: 20px;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08);
            border-left: 4px solid #667eea;
            transition: transform 0.2s, box-shadow 0.2s;
        }}

        .stat-card:hover {{
            transform: translateY(-2px);
            box-shadow: 0 4px 16px rgba(0, 0, 0, 0.12);
        }}

        .stat-card.strong {{
            border-left-color: #28a745;
        }}

        .stat-card.weak {{
            border-left-color: #ffc107;
        }}

        .stat-card.not-related {{
            border-left-color: #dc3545;
        }}

        .stat-card.empty {{
            border-left-color: #e3e0e0;
        }}

        .stat-label {{
            font-size: 0.9rem;
            color: #6c757d;
            margin-bottom: 8px;
        }}

        .stat-value {{
            font-size: 2rem;
            font-weight: 700;
            color: #212529;
        }}

        .table-section {{
            padding: 40px;
            overflow-x: auto;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.95rem;
        }}

        thead {{
            position: sticky;
            top: 0;
            z-index: 10;
        }}

        th {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            font-weight: 600;
            padding: 16px 12px;
            text-align: center;
            position: relative;
            white-space: nowrap;
        }}

        th:first-child {{
            border-top-left-radius: 8px;
        }}

        th:last-child {{
            border-top-right-radius: 8px;
        }}

        td {{
            padding: 15px 12px;
            border-bottom: 1px solid #dee2e6;
            border-right: 1px solid #dee2e6;
            vertical-align: top;
            background: white;
        }}

        td:last-child {{
            border-right: none;
        }}

        tbody tr:hover {{
            background: #f8f9fa;
        }}

        /* 条款内容列 - 加大宽度比例 */
        td.clause-a,
        td.clause-b {{
            text-align: left;
            vertical-align: top;
            font-size: 0.95rem;
            line-height: 1.8;
            max-width: 800px;
            min-width: 300px;
            white-space: pre-wrap;
            word-wrap: break-word;
        }}

        /* 路径列 - 缩小宽度比例 */
        td.path {{
            text-align: left;
            font-size: 0.85rem;
            color: #6c757d;
            font-family: "Courier New", monospace;
            max-width: 250px;
            min-width: 100px;
        }}

        /* 数值列 */
        td.score,
        td.rank {{
            text-align: center;
            font-weight: 500;
        }}

        /* 相关性标签 */
        .relevance-badge {{
            display: inline-block;
            padding: 6px 16px;
            border-radius: 20px;
            font-weight: 600;
            font-size: 0.9rem;
            text-align: center;
        }}

        .relevance-strong {{
            background: #d4edda;
            color: #155724;
        }}

        .relevance-weak {{
            background: #fff3cd;
            color: #856404;
        }}

        .relevance-not-related {{
            background: #f8d7da;
            color: #721c24;
        }}

        /* 理由列 */
        td.reason {{
            text-align: left;
            font-size: 0.9rem;
            color: #495057;
            line-height: 1.6;
            max-width: 400px;
        }}

        /* 空匹配行 */
        tr.empty-match {{
            background: #fff5f5 !important;
        }}

        tr.empty-match td {{
            color: #999;
            font-style: italic;
        }}

        /* 分隔线 */
        .divider {{
            height: 1px;
            background: linear-gradient(90deg, transparent, #dee2e6, transparent);
            margin: 20px 0;
        }}

        /* 响应式 */
        @media (max-width: 768px) {{
            .stats-grid {{
                grid-template-columns: 1fr;
            }}

            table {{
                font-size: 0.85rem;
            }}

            th, td {{
                padding: 10px 8px;
            }}
        }}

        /* 打印样式 */
        @media print {{
            body {{
                background: white;
                padding: 0;
            }}

            .container {{
                box-shadow: none;
                border-radius: 0;
            }}

            .header {{
                background: #333 !important;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        {self._generate_header_html(df)}
        {stats_html}
        {table_html}
    </div>
</body>
</html>"""
        return html

    def _generate_header_html(self, df: pd.DataFrame) -> str:
        """生成页面头部 HTML"""
        return f"""        <div class="header">
            <h1>RBA Clause Matching Results</h1>
            <div class="subtitle">
                Generation Time: {time.strftime('%Y-%m-%d %H:%M:%S')}
            </div>
        </div>"""

    def _generate_stats_html(self, stats: Dict[str, Any], doc_counts: Dict[str, int]) -> str:
        """生成统计信息 HTML"""
        relevance_counts = stats['relevance_counts']

        cards = []

        # A/B 文档数量卡片（使用 JSON 文件名）
        cards.append(f"""                <div class="stat-card">
                    <div class="stat-label">📄 {doc_counts.get('a_file_name', 'A 文档')}</div>
                    <div class="stat-value">{doc_counts.get('a_docs', 0)}</div>
                </div>""")
        cards.append(f"""                <div class="stat-card">
                    <div class="stat-label">📄 {doc_counts.get('b_file_name', 'B 文档')}</div>
                    <div class="stat-value">{doc_counts.get('b_docs', 0)}</div>
                </div>""")

        # 总数卡片
        cards.append(f"""                <div class="stat-card">
                    <div class="stat-label">总记录数</div>
                    <div class="stat-value">{stats['total']}</div>
                </div>""")

        # 空匹配卡片
        if stats['empty_match'] > 0:
            cards.append(f"""                <div class="stat-card empty">
                    <div class="stat-label">空匹配（无匹配结果）</div>
                    <div class="stat-value">{stats['empty_match']}</div>
                </div>""")

        # 相关性统计卡片（按指定顺序：强相关 > 弱相关 > 不相关）
        relevance_order = ['强相关', '弱相关', '不相关']
        for relevance in relevance_order:
            if relevance in relevance_counts:
                count = relevance_counts[relevance]
                css_class = 'strong' if relevance == '强相关' else ('weak' if relevance == '弱相关' else 'not-related')
                icon = '🟢' if relevance == '强相关' else ('🟡' if relevance == '弱相关' else '🔴')
                cards.append(f"""                <div class="stat-card {css_class}">
                    <div class="stat-label">{icon} {relevance}</div>
                    <div class="stat-value">{count}</div>
                </div>""")

        return f"""        <div class="stats-section">
            <div class="stats-title">匹配统计</div>
            <div class="stats-grid">
{chr(10).join(cards)}
            </div>
        </div>"""

    def _calculate_merge_spans(self, df: pd.DataFrame) -> dict:
        """计算需要合并的单元格的 rowspan"""
        merge_spans = {}
        start_row = 0
        current_value = None
        merge_start_row = 0

        for row_idx in range(len(df)):
            cell_value = df.iloc[row_idx]['A文件条款']

            if cell_value != current_value:
                if current_value is not None and merge_start_row < row_idx:
                    span = row_idx - merge_start_row
                    for r in range(merge_start_row, row_idx):
                        merge_spans[r] = span if r == merge_start_row else 0

                current_value = cell_value
                merge_start_row = row_idx

        # 处理最后一组
        if merge_start_row < len(df):
            span = len(df) - merge_start_row
            for r in range(merge_start_row, len(df)):
                merge_spans[r] = span if r == merge_start_row else 0

        return merge_spans

    def _generate_table_html(self, df: pd.DataFrame, columns_order: List[str], merge_spans: dict, doc_counts: Dict[str, int]) -> str:
        """生成表格 HTML"""

        # 表头（使用文件名，已去掉 .json 后缀）
        a_name = doc_counts.get('a_file_name', 'A 文档')
        b_name = doc_counts.get('b_file_name', 'B 文档')

        header_mapping = {
            'A文件条款': f'{a_name}',
            'B文件条款': f'{b_name}',
            '向量相似度': 'Vector_Score',
            'Rerank分数': 'Rerank_Score',
            'LLM判断结果': 'Relevance_label',
            'LLM判断理由': 'LLM_Rationale',
            '排名': 'Rank',
            'A文件路径': f'{a_name} clause path',
            'B文件路径': f'{b_name} clause path',
        }

        headers = [header_mapping.get(col, col) for col in columns_order]

        thead_html = "        <thead>\n            <tr>\n"
        for h in headers:
            thead_html += f"                <th>{h}</th>\n"
        thead_html += "            </tr>\n        </thead>"

        # 表体
        tbody_html = "        <tbody>\n"

        for row_idx, row in df.iterrows():
            # 判断是否为空匹配行
            is_empty = row['LLM判断结果'] == ''
            tr_class = ' class="empty-match"' if is_empty else ''

            tbody_html += f"            <tr{tr_class}>\n"

            for col_idx, col in enumerate(columns_order):
                value = row[col]

                # 处理 A 文件条款的合并单元格
                if col == 'A文件条款':
                    rowspan = merge_spans.get(row_idx, 1)
                    if rowspan == 0:
                        continue  # 跳过被合并的单元格
                    rowspan_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
                else:
                    rowspan_attr = ''

                # 确定单元格的 CSS 类
                cell_class_attr = self._get_cell_class_attr(col, value, is_empty)

                # 格式化单元格内容
                cell_content = self._format_cell_content(col, value, is_empty)

                # 拼接单元格 HTML (确保 class 前有空格)
                class_space = ' ' if cell_class_attr else ''
                tbody_html += f'                <td{class_space}{cell_class_attr}{rowspan_attr}>{cell_content}</td>\n'

            tbody_html += "            </tr>\n"

        tbody_html += "        </tbody>"

        return f"""        <div class="table-section">
            <table>
{thead_html}
{tbody_html}
            </table>
        </div>"""

    def _get_cell_class_attr(self, col: str, value: Any, is_empty: bool) -> str:
        """获取单元格的 class 属性字符串（含 class= 前缀）"""
        if is_empty:
            return ''

        class_map = {
            'A文件条款': 'class="clause-a"',
            'B文件条款': 'class="clause-b"',
            'A文件路径': 'class="path"',
            'B文件路径': 'class="path"',
            '向量相似度': 'class="score"',
            'Rerank分数': 'class="score"',
            '排名': 'class="rank"',
            'LLM判断理由': 'class="reason"',
        }
        return class_map.get(col, '')

    def _format_cell_content(self, col: str, value: Any, is_empty: bool) -> str:
        """格式化单元格内容"""
        if pd.isna(value) or value == '':
            return '<span style="color: #999;">—</span>'

        if col == 'LLM判断结果':
            css_class = ''
            if value == '强相关':
                css_class = 'relevance-strong'
            elif value == '弱相关':
                css_class = 'relevance-weak'
            elif value == '不相关':
                css_class = 'relevance-not-related'
            return f'<span class="relevance-badge {css_class}">{value}</span>'

        if col in ['向量相似度', 'Rerank分数']:
            return f'{value:.4f}'

        # HTML 转义
        if isinstance(value, str):
            value = value.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        return str(value)


# ==================== 主程序入口 ====================
def main():
    """主程序"""
    # 检查文件是否存在
    if not os.path.exists(Config.A_FILE):
        print(f"错误: 找不到文件 {Config.A_FILE}")
        return

    if not os.path.exists(Config.B_FILE):
        print(f"错误: 找不到文件 {Config.B_FILE}")
        return

    # 创建匹配器并执行匹配
    matcher = TextMatcher()
    results = matcher.match()

    # 导出结果
    if results:
        # 导出 Excel
        stats = matcher.export_to_excel(results)

        # 导出 HTML
        matcher.export_to_html(results, stats, matcher.doc_counts)

        # 打印统计摘要
        print("\n" + "=" * 60)
        print("导出完成！")
        print(f"  - Excel: {Config.OUTPUT_EXCEL}")
        print(f"  - HTML:  {Config.OUTPUT_HTML}")
        print("=" * 60)

        # 打印统计信息
        print("\n结果统计:")
        print(f"  - 总记录数: {stats['total']}")
        print(f"  - 空匹配（无匹配结果）: {stats['empty_match']} 条")
        for relevance, count in stats['relevance_counts'].items():
            print(f"  - {relevance}: {count} 条")

    else:
        print("\n没有找到匹配结果")


if __name__ == "__main__":
    main()
