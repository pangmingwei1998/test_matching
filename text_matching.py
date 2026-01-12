#!/usr/bin/env python3
"""
责任标准文档自动化比对脚本
功能：语义向量召回 + LLM精判 + Excel导出
"""

import json
import os
import time
from typing import List, Dict, Tuple, Any
import requests
import numpy as np
import pandas as pd
import openai
from tqdm import tqdm

import faiss
import torch
from transformers import AutoTokenizer, AutoModel


# ==================== 配置参数 ====================
class Config:
    # 文件路径
    ALLIANCE_FILE = "/home/pmw/h20/Text_matching/企业联盟责任标准.json"
    SUPPLIER_FILE = "/home/pmw/h20/Text_matching/供应商责任标准.json"
    OUTPUT_EXCEL = "/home/pmw/h20/Text_matching/matching_results.xlsx"

    # BGE-M3 模型
    # 支持本地路径或 Hugging Face 模型名
    EMBEDDING_MODEL = "BAAI/bge-m3"
    EMBEDDING_DEVICE = "cuda" if torch.cuda.is_available() else "cpu"

    # 检查本地缓存是否存在
    HF_CACHE = os.path.expanduser("~/.cache/huggingface/hub/models--BAAI--bge-m3")
    USE_LOCAL_ONLY = os.path.exists(HF_CACHE)  # 如果缓存存在，强制使用本地模式

    # 向量检索参数
    TOP_K = 5  # 召回Top-K候选

    # LLM API 配置
    LLM_API_BASE = "http://10.71.5.24:8000/v1"
    LLM_API_KEY = "empty"  # 本地服务通常不需要key
    LLM_MODEL = "gpt-3.5-turbo"
    LLM_TIMEOUT = 60
    LLM_MAX_RETRIES = 3

    # 相似度阈值
    SIMILARITY_THRESHOLD = 0.8  # 低于此分数的匹配不进行LLM判断


# ==================== 工具函数 ====================
def load_json_documents(file_path: str) -> List[Dict[str, Any]]:
    """加载JSON格式的文档，支持多行JSON对象"""
    documents = []

    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # 尝试方式1: 每行一个JSON对象
    lines = content.split('\n')
    line_num = 0
    while line_num < len(lines):
        line = lines[line_num].strip()
        if not line:
            line_num += 1
            continue

        try:
            # 尝试单行解析
            doc = json.loads(line)
            doc['_line_num'] = line_num + 1
            documents.append(doc)
            line_num += 1
        except json.JSONDecodeError:
            # 尝试多行解析（找到完整的JSON对象）
            json_str = line
            nested_line = line_num + 1
            while nested_line < len(lines):
                next_line = lines[nested_line]
                json_str += '\n' + next_line
                try:
                    doc = json.loads(json_str)
                    doc['_line_num'] = line_num + 1
                    documents.append(doc)
                    line_num = nested_line + 1
                    break
                except json.JSONDecodeError:
                    nested_line += 1
            else:
                print(f"警告: 第{line_num + 1}行JSON解析失败（尝试了多行合并）")
                line_num += 1

    return documents


def truncate_text(text: str, max_length: int = 512) -> str:
    """截断过长的文本"""
    if len(text) <= max_length:
        return text
    return text[:max_length] + "..."


# ==================== BGE-M3 向量嵌入 ====================
class BGEEmbedder:
    """使用 BGE-M3 模型生成文本嵌入向量"""

    def __init__(self, model_name: str = Config.EMBEDDING_MODEL, device: str = Config.EMBEDDING_DEVICE):
        print(f"正在加载 BGE-M3 模型 ({device}模式)...")
        self.device = device

        # 强制使用离线模式，禁用任何网络请求
        os.environ['HF_HUB_OFFLINE'] = '1'
        os.environ['TRANSFORMERS_OFFLINE'] = '1'
        # 禁用 huggingface_hub 的网络检查
        os.environ['HF_HUB_DISABLE_TELEMETRY'] = '1'

        # 获取本地模型路径
        local_model_path = None
        if Config.USE_LOCAL_ONLY:
            # 查找实际的 snapshot 路径
            import glob
            snapshot_pattern = os.path.expanduser("~/.cache/huggingface/hub/models--BAAI--bge-m3/snapshots/*")
            snapshot_dirs = glob.glob(snapshot_pattern)
            if snapshot_dirs:
                # 找到包含完整模型文件的 snapshot（需要有 config.json 或 tokenizer.json）
                valid_snapshot = None
                for snapshot_dir in snapshot_dirs:
                    if os.path.exists(os.path.join(snapshot_dir, "config.json")) and \
                       os.path.exists(os.path.join(snapshot_dir, "tokenizer.json")):
                        valid_snapshot = snapshot_dir
                        break

                if valid_snapshot:
                    print(f"  使用本地缓存: {valid_snapshot}")
                    model_to_load = valid_snapshot
                else:
                    print(f"  本地缓存不完整，尝试使用模型名: {model_name}")
                    model_to_load = model_name
            else:
                print(f"  本地缓存未找到，尝试使用模型名: {model_name}")
                model_to_load = model_name
        else:
            print(f"  使用远程模型: {model_name}")
            print("  提示: 首次运行会下载模型（约2GB）")
            model_to_load = model_name

        load_kwargs = {
            "local_files_only": True,
            "trust_remote_code": True
        }

        try:
            self.tokenizer = AutoTokenizer.from_pretrained(model_to_load, **load_kwargs)
            self.model = AutoModel.from_pretrained(model_to_load, **load_kwargs)
            self.model.to(device)
            self.model.eval()
            print("BGE-M3 模型加载完成")

        except Exception as e:
            print(f"\n错误: 模型加载失败: {e}")
            if Config.USE_LOCAL_ONLY:
                print("\n本地缓存可能损坏，请尝试:")
                print("  rm -rf ~/.cache/huggingface/hub/models--BAAI--bge-m3")
                print("  然后重新运行脚本")
            else:
                print("\n解决方案:")
                print("1. 等待一段时间后重试")
                print("2. 设置环境变量使用镜像站: export HF_ENDPOINT=https://hf-mirror.com")
            raise

    def encode(self, texts: List[str], batch_size: int = 32) -> np.ndarray:
        """将文本列表编码为向量"""
        all_embeddings = []

        for i in range(0, len(texts), batch_size):
            batch_texts = texts[i:i + batch_size]

            # Tokenize
            encoded_input = self.tokenizer(
                batch_texts,
                padding=True,
                truncation=True,
                max_length=512,
                return_tensors='pt'
            )

            # Move to device
            encoded_input = {k: v.to(self.device) for k, v in encoded_input.items()}

            # Encode
            with torch.no_grad():
                model_output = self.model(**encoded_input)
                # 使用 [CLS] token 的嵌入或者平均池化
                embeddings = self._mean_pooling(model_output, encoded_input['attention_mask'])
                # 归一化
                embeddings = torch.nn.functional.normalize(embeddings, p=2, dim=1)

            all_embeddings.append(embeddings.cpu().numpy())

        return np.vstack(all_embeddings)

    def _mean_pooling(self, model_output, attention_mask):
        """平均池化"""
        token_embeddings = model_output[0]
        input_mask_expanded = attention_mask.unsqueeze(-1).expand(token_embeddings.size()).float()
        return torch.sum(token_embeddings * input_mask_expanded, 1) / torch.clamp(input_mask_expanded.sum(1), min=1e-9)


# ==================== 向量索引 ====================
class VectorIndex:
    """使用 FAISS 构建向量索引"""

    def __init__(self, embeddings: np.ndarray):
        dimension = embeddings.shape[1]
        self.index = faiss.IndexFlatIP(dimension)  # 内积相似度（向量已归一化）
        self.index.add(embeddings.astype('float32'))
        self.dimension = dimension

    def search(self, query_embedding: np.ndarray, top_k: int = 5) -> Tuple[np.ndarray, np.ndarray]:
        """搜索最相似的向量
        返回: (相似度分数, 索引)
        """
        similarities, indices = self.index.search(query_embedding.astype('float32'), top_k)
        return similarities, indices


# ==================== LLM 精判 ====================
class LLMJudge:
    """使用 LLM 判断两个段落的相关性"""

    # 相关性等级
    RELEVANCE_NOT_RELATED = "不相关"
    RELEVANCE_WEAK = "弱相关"
    RELEVANCE_STRONG = "强相关"

    def __init__(self):
        self.client = openai.OpenAI(
            api_key=Config.LLM_API_KEY,
            base_url=Config.LLM_API_BASE
        )

    def judge(self, text1: str, text2: str) -> Tuple[str, str]:
        """判断两段文本的相关性
        返回: (相关性等级, 理由说明)
        """
        prompt = self._build_prompt(text1, text2)

        for attempt in range(Config.LLM_MAX_RETRIES):
            try:
                response = self.client.chat.completions.create(
                    model=Config.LLM_MODEL,
                    messages=[
                        {
                            "role": "system",
                            "content": "你是一个专业的责任标准文档分析专家。你需要判断两段文本在'责任义务层面'是否相关。"
                        },
                        {
                            "role": "user",
                            "content": prompt
                        }
                    ],
                    temperature=0.1,
                    timeout=Config.LLM_TIMEOUT
                )

                result = response.choices[0].message.content.strip()
                return self._parse_result(result)

            except Exception as e:
                if attempt < Config.LLM_MAX_RETRIES - 1:
                    wait_time = (attempt + 1) * 2
                    print(f"LLM 调用失败，{wait_time}秒后重试... 错误: {e}")
                    time.sleep(wait_time)
                else:
                    print(f"LLM 调用失败，使用默认值: {e}")
                    return self.RELEVANCE_NOT_RELATED, f"调用失败: {str(e)}"

    def _build_prompt(self, text1: str, text2: str) -> str:
        """构建 LLM 判断提示词"""
        return f"""请判断以下两段责任标准条款在"责任义务层面"是否相关。

【条款 A】（企业联盟责任标准）：
{text1}

【条款 B】（供应商责任标准）：
{text2}

请从以下几个方面判断：
1. 是否涉及相似的责任或义务主题
2. 是否规定相似的要求或标准
3. 是否针对相似的利益相关方

请仅返回以下格式的结果（不要输出其他内容）：
相关性：[不相关/弱相关/强相关]
理由：[简要说明判断理由，不超过100字]
"""

    def _parse_result(self, result: str) -> Tuple[str, str]:
        """解析 LLM 返回结果"""
        result = result.strip()

        # 提取相关性等级
        relevance = self.RELEVANCE_NOT_RELATED
        if self.RELEVANCE_STRONG in result:
            relevance = self.RELEVANCE_STRONG
        elif self.RELEVANCE_WEAK in result:
            relevance = self.RELEVANCE_WEAK
        elif self.RELEVANCE_NOT_RELATED in result:
            relevance = self.RELEVANCE_NOT_RELATED

        # 提取理由
        reason = ""
        if "理由：" in result:
            reason = result.split("理由：", 1)[1].strip()
        elif "Reason:" in result:
            reason = result.split("Reason:", 1)[1].strip()

        return relevance, reason


# ==================== 主匹配流程 ====================
class TextMatcher:
    """文本匹配主流程"""

    def __init__(self):
        print("=" * 60)
        print("责任标准文档自动化比对系统")
        print("=" * 60)

        # 加载文档
        print("\n[1/4] 加载文档...")
        self.alliance_docs = load_json_documents(Config.ALLIANCE_FILE)
        self.supplier_docs = load_json_documents(Config.SUPPLIER_FILE)
        print(f"  - 企业联盟责任标准: {len(self.alliance_docs)} 条")
        print(f"  - 供应商责任标准: {len(self.supplier_docs)} 条")

        # 初始化嵌入模型
        print("\n[2/4] 初始化 BGE-M3 嵌入模型...")
        self.embedder = BGEEmbedder()

        # 构建供应商文档向量索引
        print("\n[3/4] 构建供应商文档向量索引...")
        supplier_texts = [doc.get('text', '') for doc in self.supplier_docs]
        supplier_embeddings = self.embedder.encode(supplier_texts)
        self.vector_index = VectorIndex(supplier_embeddings)
        print(f"  - 向量维度: {supplier_embeddings.shape[1]}")
        print(f"  - 索引完成")

        # 初始化 LLM 判断器
        print("\n[4/4] 初始化 LLM 判断器...")
        self.llm_judge = LLMJudge()
        print("  - API 地址:", Config.LLM_API_BASE)
        print("  - 模型:", Config.LLM_MODEL)

        print("\n" + "=" * 60)
        print("初始化完成，开始匹配...")
        print("=" * 60 + "\n")

    def match(self) -> List[Dict[str, Any]]:
        """执行匹配流程"""
        results = []

        # 对企业联盟的每个段落进行匹配
        for alliance_doc in tqdm(self.alliance_docs, desc="匹配进度"):
            alliance_text = alliance_doc.get('text', '')

            # 1. 向量检索 Top-K
            query_embedding = self.embedder.encode([alliance_text])
            similarities, indices = self.vector_index.search(query_embedding, Config.TOP_K)

            # 2. 对 Top-K 结果进行 LLM 精判
            for rank, (similarity, idx) in enumerate(zip(similarities[0], indices[0]), 1):
                supplier_doc = self.supplier_docs[idx]
                supplier_text = supplier_doc.get('text', '')

                # 相似度过滤
                if similarity < Config.SIMILARITY_THRESHOLD:
                    continue

                # LLM 精判
                llm_relevance, llm_reason = self.llm_judge.judge(alliance_text, supplier_text)

                # 保存结果（使用完整文本，不截断）
                result = {
                    '企业联盟条款': alliance_text,
                    '供应商条款': supplier_text,
                    '相似度得分': round(float(similarity), 4),
                    '排名': rank,
                    'LLM判断结果': llm_relevance,
                    'LLM判断理由': llm_reason,
                    '供应商标题': supplier_doc.get('title', ''),
                    '企业版本': alliance_doc.get('Version', ''),
                }
                results.append(result)

        return results

    def export_to_excel(self, results: List[Dict[str, Any]], output_path: str = None):
        """导出结果到 Excel，相同的"企业联盟条款"合并单元格"""
        if output_path is None:
            output_path = Config.OUTPUT_EXCEL

        print(f"\n正在导出结果到 {output_path}...")

        df = pd.DataFrame(results)

        # 调整列顺序
        columns_order = [
            '企业联盟条款',
            '供应商条款',
            '相似度得分',
            'LLM判断结果',
            'LLM判断理由',
            '排名',
            '供应商标题',
            '企业版本',
        ]

        # 只保留存在的列
        columns_order = [col for col in columns_order if col in df.columns]
        df = df[columns_order]

        # 导出到 Excel
        from openpyxl.styles import Alignment, Font, Border, Side
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='匹配结果')
            worksheet = writer.sheets['匹配结果']

            # 调整列宽
            worksheet.column_dimensions['A'].width = 60  # 企业联盟条款
            worksheet.column_dimensions['B'].width = 60  # 供应商条款
            worksheet.column_dimensions['C'].width = 15  # 相似度得分
            worksheet.column_dimensions['D'].width = 15  # LLM判断结果
            worksheet.column_dimensions['E'].width = 40  # LLM判断理由
            worksheet.column_dimensions['F'].width = 10  # 排名
            worksheet.column_dimensions['G'].width = 30  # 供应商标题
            worksheet.column_dimensions['H'].width = 15  # 企业版本

            # 设置所有数据行的行高为 200
            for row in range(2, len(df) + 2):  # 从第2行开始（第1行是标题）
                worksheet.row_dimensions[row].height = 200

            # 设置标题行样式
            header_font = Font(bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for cell in worksheet[1]:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # 合并相同的企业联盟条款单元格
            # 从数据行开始（第2行，Excel索引为2）
            start_row = 2
            current_value = None
            merge_start_row = 2

            for row_idx in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_idx, column=1).value

                if cell_value != current_value:
                    # 如果之前的值相同且有多行，合并单元格
                    if current_value is not None and merge_start_row < row_idx - 1:
                        worksheet.merge_cells(f'A{merge_start_row}:A{row_idx - 1}')
                        # 设置合并后的单元格对齐方式
                        merged_cell = worksheet.cell(row=merge_start_row, column=1)
                        merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    current_value = cell_value
                    merge_start_row = row_idx

            # 处理最后一组相同的值
            if merge_start_row < len(df) + 2:
                worksheet.merge_cells(f'A{merge_start_row}:A{len(df) + 1}')
                merged_cell = worksheet.cell(row=merge_start_row, column=1)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # 设置所有数据行的对齐方式和边框
            for row in range(2, len(df) + 2):
                for col in range(1, len(columns_order) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = Alignment(
                        horizontal='left' if col in [1, 2, 5, 7] else 'center',
                        vertical='top',
                        wrap_text=True
                    )
                    cell.border = thin_border

        print(f"导出完成！共 {len(results)} 条匹配结果")

        # 输出统计信息
        if results:
            print("\n结果统计:")
            relevance_counts = {}
            for r in results:
                relevance = r['LLM判断结果']
                relevance_counts[relevance] = relevance_counts.get(relevance, 0) + 1

            for relevance, count in relevance_counts.items():
                print(f"  - {relevance}: {count} 条")


# ==================== 主程序入口 ====================
def main():
    """主程序"""
    # 检查文件是否存在
    if not os.path.exists(Config.ALLIANCE_FILE):
        print(f"错误: 找不到文件 {Config.ALLIANCE_FILE}")
        return

    if not os.path.exists(Config.SUPPLIER_FILE):
        print(f"错误: 找不到文件 {Config.SUPPLIER_FILE}")
        return

    # 创建匹配器并执行匹配
    matcher = TextMatcher()
    results = matcher.match()

    # 导出结果
    if results:
        matcher.export_to_excel(results)
    else:
        print("\n没有找到匹配结果")


if __name__ == "__main__":
    main()
